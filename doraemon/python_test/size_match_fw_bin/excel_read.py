#!/usr/bin/python
from openpyxl import load_workbook


list_3te7 = []
list_0721_cell = []
list_1125_cell = []

sn_0721_list = []
sn_1125_list = []

wb = load_workbook('3TE7_BiCS4_record.xlsx')
ws1 = wb.create_sheet("0721_sn_list")
ws2 = wb.create_sheet("1125_sn_list")
sheet_0 = wb['20200721']
sheet_1 = wb['20201125']

def get_sheet_0721_data():
    for row in sheet_0.iter_rows(min_row=2, max_col=2):
        list_0721_cell.append([cell.value for cell in row])

    for line in list_0721_cell:
        if line in sn_0721_list:
            continue
        elif "02T" in line[1]:
            print(line)
        else:
            sn_0721_list.append(line)

    for sn_num in sn_0721_list:
        if "64G" in sn_num[1]:
            sn_num.append("Q2TNCGAC.bin")
        elif "A28" in sn_num[1]:
            sn_num.append("Q2TNCGBC.bin")
        elif "B56" in sn_num[1] and "QL" in line[1]:
            sn_num.append("Q2TNCGDC.bin")
        elif "B56" in sn_num[1] and "DL" in line[1]:
            sn_num.append("Q2TNCGFC.bin")
        elif "C12" in sn_num[1]:
            sn_num.append("Q2TNCGHC.bin")
        elif "01T" in sn_num[1]:
            sn_num.append("Q2TNCGPC.bin")
        else:
            print("Size is not in rule")

    for row in sn_0721_list:
        ws1.append(row)
    

def get_sheet_1125_data():
    for row in sheet_1.iter_rows(min_row=2, max_col=2):
        sn_1125_list.append([cell.value for cell in row])

    for line in sn_1125_list:
        if line in sn_1125_list:
            continue
        else:
            sn_0721_list.append(line)

    for sn_num in sn_1125_list:
        if "B56" in sn_num[1]:
            sn_num.append("Q2TNCGDC_1125.bin")
        elif "C12" in sn_num[1]:
            sn_num.append("Q2TNCGHC_1125.bin")
        else:
            print("Size is not in rule")

    for row in sn_1125_list:
        ws2.append(row)


def main():
    get_sheet_0721_data()
    get_sheet_1125_data()
    wb.save(filename="test.xlsx")

if __name__ == "__main__":
    main()
